# -*- coding: utf-8 -*-
"""
Created on Fri Jul 16 09:04:59 2021

@author: Utente
"""
import openpyxl

#takes the filename as arguments and returns a workbook value
wb = openpyxl.load_workbook('BC04_54.xlsx')
sheet = wb.get_active_sheet()
#define the boundaries of the sheet
col = sheet.max_column
row = sheet.max_row

b=0

#nested for loop to establish the possible regularity
#of the analyzed values
for i in range(1, col+1):
    a1=sheet.cell(row=1, column=i)
    
    for j in range(1, row+1 ):
      if sheet.cell(row=j, column=i).value != a1.value:
        b=1
        break
      else:
        b=0
    if b == 0:
        print(a1.value)
    else:
        print('x')
        